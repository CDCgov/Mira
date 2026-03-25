# Import async packages
import asyncio

# Import shiny packages
from shiny import *

# Import python packages
import re
import os
import sys
import yaml
import signal
import requests
import subprocess
import colorlover
import pandas as pd
from datetime import datetime

# Import specific modules from Python packages
from glob import glob
from pathlib import Path
from os.path import dirname, realpath 

# Python package to parse xlsx worksheet
import openpyxl as xl
from openpyxl.worksheet.datavalidation import DataValidation as DV

# Get samplesheet column names btw ONT vs ILLUMINA
from utils.global_var import *

# Get time stamp
def timestamp() -> str:
    return datetime.now().strftime("%x %X")
      
# Get current version of mira
def get_config(config_file):
    try:
        with open(config_file, "r") as y:
            config = yaml.safe_load(y)
        return config
    except Exception as e:
        print(f"ERROR: Invalid config_file at {config_file}.\nException:{e}", file=sys.stderr)

# Determine whether the app is in dev or prod
def get_version():
    descript_dict = {}
    description_file = f"{os.path.dirname(os.path.realpath(__file__))}/../DESCRIPTION"
    with open(description_file, 'r') as infi:
        for line in infi:
            try:
                descript_dict[line.split(':')[0]]=line.split(":")[1]
            except:
                continue
    # Make sure version in dict
    if isinstance(descript_dict, dict) and "Version" in descript_dict.keys():
        return descript_dict['Version'].strip()

# Create a modal to display announcements of whenever the new version of MIRA is available
def check_deployment_version(config):
    # Get deployment version
    VERSION_URL = config["VERSION_URL"]
    # Get current version
    current_version = get_version()
    # Get available version on Github       
    github_version = requests.get(VERSION_URL)
    available_version = re.findall(r"Version.+(?=\n)", github_version.text)[0]
    # Check if current version is lesser than available version online
    if current_version >= available_version:
        # Show modal
        return ui.modal_show(
            ui.modal(
                ui.tags.div(
                    ui.tags.span(f"Current {current_version}"),
                    ui.tags.br(),
                    ui.tags.span(f"Available {available_version}"),
                ),
                ui.tags.br(),
                ui.HTML('<svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" fill="currentColor" class="bi bi-hand-index-thumb" viewBox="0 0 16 16"><path d="M6.75 1a.75.75 0 0 1 .75.75V8a.5.5 0 0 0 1 0V5.467l.086-.004c.317-.012.637-.008.816.027.134.027.294.096.448.182.077.042.15.147.15.314V8a.5.5 0 0 0 1 0V6.435l.106-.01c.316-.024.584-.01.708.04.118.046.3.207.486.43.081.096.15.19.2.259V8.5a.5.5 0 1 0 1 0v-1h.342a1 1 0 0 1 .995 1.1l-.271 2.715a2.5 2.5 0 0 1-.317.991l-1.395 2.442a.5.5 0 0 1-.434.252H6.118a.5.5 0 0 1-.447-.276l-1.232-2.465-2.512-4.185a.517.517 0 0 1 .809-.631l2.41 2.41A.5.5 0 0 0 6 9.5V1.75A.75.75 0 0 1 6.75 1M8.5 4.466V1.75a1.75 1.75 0 1 0-3.5 0v6.543L3.443 6.736A1.517 1.517 0 0 0 1.07 8.588l2.491 4.153 1.215 2.43A1.5 1.5 0 0 0 6.118 16h6.302a1.5 1.5 0 0 0 1.302-.756l1.395-2.441a3.5 3.5 0 0 0 .444-1.389l.271-2.715a2 2 0 0 0-1.99-2.199h-.581a5 5 0 0 0-.195-.248c-.191-.229-.51-.568-.88-.716-.364-.146-.846-.132-1.158-.108l-.132.012a1.26 1.26 0 0 0-.56-.642 2.6 2.6 0 0 0-.738-.288c-.31-.062-.739-.058-1.05-.046zm2.094 2.025"/></svg> '), 
                ui.tags.a(
                    "CLICK HERE TO SEE HOW TO UPGRADE MIRA TO LATEST VERSION",
                    href="https://cdcgov.github.io/MIRA/articles/upgrading-mira.html",
                    target="_blank",
                ),            
                title="A new version of MIRA is available!",
                size="l",
                easy_close=True,
                footer=ui.modal_button("Close"),
            )
        )
        
# Create a modal to display announcements, new MIRA features, and so forth
def show_info_modal():
    return ui.modal_show(
        ui.modal(
           ui.tags.a(
                "CLICK HERE TO SEE HOW TO UPGRADE MIRA TO LATEST VERSION",
                href="https://cdcgov.github.io/MIRA/articles/upgrading-mira.html",
                target="_blank",
            ),            
            title="A new version of MIRA is available!",
            size="l",
            easy_close=True,
            footer=ui.modal_button("Close"),
        )
    )

# Generate samplesheet file        
def generate_samplesheet_xl(data_root, seq_run, experiment_type):
    wb = xl.Workbook()
    ws = wb.active
    # Store possible drop down options
    for r, t in enumerate(sample_type_options):
        ws[f"Z{r+1}"].value = t
    sample_types = DV(type="list", formula1=f"=Z$1:Z$3")
    ws.add_data_validation(sample_types)
    # Check experimental type
    if experiment_type.upper() in "ONT":
        # Check fastq_pass
        bar_nums = [int(i[-2:]) for i in glob(f"{data_root}/{seq_run}/fastq_pass/b*")]
        # If barcode was not given, provide first 3 samples
        if len(bar_nums) == 0:
            bar_nums = list(range(1, 4))
        # Populate worksheet
        ws["A1"].value, ws["B1"].value, ws["C1"].value = (
            "barcode",
            "sample_id",
            "sample_type",
        )
        bar_nums.sort()
        barcodes = [f"barcode{i:02}" for i in bar_nums]
        row = 2
        for r, b in enumerate(barcodes):
            ws[f"A{row}"].value = b
            ws[f"C{row}"].value = "Test"
            row += 1
        sample_types.add(f"C2:C{len(barcodes)+1}")
    elif experiment_type.upper() in "ILLUMINA":
        # Check fastq
        fqs = [
            i.split("/")[-1]
            for i in glob(f"{data_root}/{seq_run}/fastq*/*R[12]*.fastq*")
        ]
        # Make sure samples have both R1 and R2
        if len(fqs) > 0:
            all_samples = list(set([re.findall(r".+(?=[_/.]R[12])", i)[0] for i in fqs]))
            ill_samples = [sample for sample in all_samples if len(glob(f"{data_root}/{seq_run}/fastq*/{sample}*R[12]*fastq*")) == 2] 
            if len(ill_samples) == 0:
                ill_samples = ["<sample_1>", "<sample_2>", "<sample_3>"]
        else:
            ill_samples = ["<sample_1>", "<sample_2>", "<sample_3>"]
        ws["A1"].value, ws["B1"].value = "sample_id", "sample_type"
        ill_samples.sort()
        row = 2
        for r, s in enumerate(ill_samples):
            ws[f"A{row}"].value = s
            ws[f"B{row}"].value = "Test"
            row += 1
        sample_types.add(f"B2:B{len(ill_samples)+1}")
    # Define and save samplesheet file
    samplesheet_file = f"{data_root}/{seq_run}/{seq_run}_samplesheet.xlsx"
    wb.save(samplesheet_file)
    return samplesheet_file     
  
# Parse samplesheet file
def parse_samplesheet(samplesheet_file):
    # Check file extension
    if os.path.isfile(samplesheet_file) and Path(samplesheet_file).suffix == ".csv":
        ss_df = pd.read_csv(samplesheet_file, skip_blank_lines=True)
    elif os.path.isfile(samplesheet_file) and Path(samplesheet_file).suffix in [".xls", ".xlsx"]:
        ss_df = pd.read_excel(samplesheet_file, engine="openpyxl")
    else:
        ss_df = pd.DataFrame()
    # Check experiment type
    if not ss_df.empty:
        if all([col in ss_df.columns for col in ont_ss_colnames]):
            ss_df = ss_df.loc[:, ont_ss_colnames]
            experiment_type = "ONT"
        elif all([col in ss_df.columns for col in illumina_ss_colnames]):
            ss_df = ss_df.loc[:, illumina_ss_colnames]
            experiment_type = "Illumina"
        else:
            ss_df = pd.DataFrame()
            experiment_type = ""
    else:
        experiment_type = ""
    # Remove all nas
    ss_df = ss_df.dropna(how='all')
    # Return samplesheet and experiment type
    return ss_df, experiment_type
  
# Create background color for IRMA summary table
def fill_irma_summary_tbl(df, n_bins=8, columns="all"):
    styled_df = df.copy().astype(str)
    bounds = [i * (1.0 / n_bins) for i in range(n_bins + 1)]
    if columns == "all":
        if "id" in df:
            df_numeric_columns = df.select_dtypes("number").drop(["id"], axis=1)
        else:
            df_numeric_columns = df.select_dtypes("number")
    else:
        df_noUndetermined = df[df["Sample"] != "Undetermined"]
        df_numeric_columns = df_noUndetermined[columns]
    styles = []
    ranges = {}
    for column in df_numeric_columns:
        df_max = df_numeric_columns[column].max()
        df_min = df_numeric_columns[column].min()
        ranges[column] = [((df_max - df_min) * i) + df_min for i in bounds]
        for i in range(1, len(bounds)):
            min_bound = ranges[column][i - 1]
            max_bound = ranges[column][i]
            backgroundColor = colorlover.scales[str(n_bins)]["seq"]["PuBuGn"][i - 1]
            color = "white" if i > len(bounds) / 2.0 else "inherit"
            cols = df.columns.get_loc(column)
            if i < (len(bounds) - 1):
                rows = df.query(f'`{column}` >= {min_bound}').index.tolist()
            else:
                rows = df.query(f'`{column}` >= {min_bound} and `{column}` < {max_bound}').index.tolist()
            styled_df.iloc[rows, cols] = [f'<div style="width: 100%; color: {color}; background-color: {backgroundColor};">' + str(val) + '</div>' for val in df.iloc[rows, cols].tolist()]
    # Return df with styles
    return styled_df
  
# Stop IRMA process
async def stop_irma_process(process):
    process.send_signal(signal.SIGINT)
    await process.wait()
    
# Kill IRMA process
async def kill_irma_process(process):
    process.terminate()
    await process.wait()   

# Force kill IRMA process
async def force_kill_irma_process(process):
    os.kill(process.pid, signal.SIGKILL)
    await process.wait() 
